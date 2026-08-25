"""Read a Bulk Mentions export (xlsx/csv), fill it from fetched Reddit data,
and write the output workbook — all in memory (bytes in, bytes out).

xlsx reads go through python-calamine and xlsx writes through xlsxwriter's
constant-memory mode — NOT openpyxl's normal mutable-cell mode, which is
catastrophically slow/memory-hungry for wide real-world exports. Measured on
a real 73,384-row x 193-column Brandwatch export: openpyxl's normal
`load_workbook()` took ~220s and ~5.7GB RAM; python-calamine read the same
file in ~12s at ~978MB, and xlsxwriter's constant_memory mode wrote it back
out in ~17s — verified byte-for-byte round-trip identical.

xlsxwriter needs `strings_to_urls=False` — by default it auto-detects
URL-looking strings and converts them to real hyperlinks, and Excel caps a
sheet at 65,530 hyperlinks; past that it silently drops the cell rather
than erroring. A Url column alone exceeds that on a file this size.

BUT python-calamine is not actually lazy: its Rust core parses and holds the
*entire* sheet in memory the instant you open it (`get_sheet_by_index()`),
before `.to_python()`/`.iter_rows()` is even called — so memory scales with
total cell count regardless of which of its own APIs you use afterward.
Measured: ~978MB for a 73k x 193 file (~14.2M cells), ~2GB for a 203k x 194
file (~39.5M cells) — the latter OOM-crashes a ~1GB-RAM hosted container
outright (Streamlit's own generic crash screen, not a normal exception).

`_predict_needs_streaming` peeks at the xlsx zip's central directory (free —
no decompression) to estimate the decompressed sheet-XML size, which predicts
calamine's peak RSS at roughly a 1.2-1.3x multiplier. Past a threshold, we
switch to `_stream_xlsx_rows` (openpyxl `read_only=True`), which IS genuinely
lazy — flat ~85MB regardless of row count — at the cost of speed (~165s to
scan 203k rows vs. ~25s with calamine). `parse_export`/`fill_export` use this
for a real two-pass streaming design (pass 1: scan for URLs only, never
holding row data; pass 2: re-stream + transform + write one row at a time)
so peak memory for Mention Filler stays flat regardless of file size.

This does NOT extend to the enrichment modules' `Sheet` class below, which
still needs the whole matrix in memory for random-access column writes (and
Theme Summary/Driver Analysis need a full-corpus view before tagging any
row) — `load_sheet_for_enrichment`/`read_header`/`iter_column_values` get the
same reader swap on the read step (roughly halves peak memory for a huge
already-filled file) but that's a partial mitigation, not the same
any-size guarantee Mention Filler gets.
"""
import csv
import datetime
import io
import itertools
import re
import zipfile
from collections import defaultdict

import openpyxl
import python_calamine
import xlsxwriter

from .text_utils import ILLEGAL, EXCEL_CELL_LIMIT, EXTRA_COLS, build_extra_cols, categorize_status, md_to_text, norm, parse_score, row_type

REQUIRED_COLS = ("Date", "Url", "Full Text", "Author", "Title")

# Decompressed-content-size proxy (xlsx: sheet+sharedStrings XML; csv: raw
# bytes) above which we route through the slow-but-flat-memory streaming
# readers instead of python-calamine. Calibrated against two real files: a
# 767MB-XML file that used ~978MB RSS via calamine (fine on the hosted app)
# and a 1,712MB-XML file that used ~2GB RSS (OOM-crashed it).
STREAMING_SIZE_THRESHOLD = 900_000_000


class BadExport(Exception):
    pass


def _predict_needs_streaming(file_bytes, is_csv):
    """Cheap, pre-parse guess at whether this file is large enough that
    python-calamine's full-sheet parse risks OOMing a ~1GB-RAM container.
    Any failure to inspect falls back to False (the existing, proven path)
    rather than guessing wrong in the riskier direction."""
    if is_csv:
        return len(file_bytes) > STREAMING_SIZE_THRESHOLD
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            names = z.namelist()
            total = sum(z.getinfo(n).file_size for n in names
                        if n.startswith("xl/worksheets/") and n.endswith(".xml"))
            total += sum(z.getinfo(n).file_size for n in names if "sharedStrings" in n)
        return total > STREAMING_SIZE_THRESHOLD
    except Exception:
        return False


def _read_xlsx_rows(file_bytes):
    """Every row as plain Python lists, via python-calamine — fast, but see
    module docstring for why it's only safe below STREAMING_SIZE_THRESHOLD."""
    wb = python_calamine.CalamineWorkbook.from_filelike(io.BytesIO(file_bytes))
    sheet = wb.get_sheet_by_index(0)
    return sheet.to_python()


def _read_csv_rows(file_bytes):
    csv.field_size_limit(50_000_000)
    text = file_bytes.decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


def _stream_xlsx_rows(file_bytes):
    """Row-by-row generator via openpyxl's read_only mode — genuinely lazy
    (flat memory regardless of row count), unlike python-calamine. Slower;
    use only when _predict_needs_streaming said the fast path is risky.

    Normalizes blank cells to "" (openpyxl yields None; calamine's
    to_python() yields "") so the two readers are interchangeable for every
    downstream consumer — confirmed by a byte-for-byte diff against the fast
    path on a real file."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        for row in wb.active.iter_rows(values_only=True):
            yield ["" if c is None else c for c in row]
    finally:
        wb.close()


def _stream_csv_rows(file_bytes):
    """Like _read_csv_rows but never materializes the whole decoded text or
    row list — csv.reader is already lazy; this just doesn't defeat that."""
    csv.field_size_limit(50_000_000)
    text_stream = io.TextIOWrapper(io.BytesIO(file_bytes), encoding="utf-8-sig", newline="")
    yield from csv.reader(text_stream)


def locate_header(rows):
    """Find the 'Query Id' header row -> (index, col_indexes). Works with or without preamble."""
    hdr_i = next((i for i, r in enumerate(rows[:30]) if r and r[0] == "Query Id"), None)
    if hdr_i is None:
        raise BadExport("Could not find the 'Query Id' header row — is this a Bulk Mentions export?")
    header = [c if c is not None else "" for c in rows[hdr_i]]
    try:
        cols = {name: header.index(name) for name in REQUIRED_COLS}
    except ValueError as e:
        raise BadExport(f"Missing expected column: {e}")
    return hdr_i, cols


def _locate_header_streaming(row_iter):
    """Like locate_header, but consumes rows from a generator one at a time
    instead of indexing into a materialized list — bounded by the same
    30-row window. Returns (preamble, header, cols); `row_iter` is left
    positioned right after the header row so the caller can keep iterating
    it for the data rows."""
    buffer = []
    for row in row_iter:
        buffer.append(row)
        if row and row[0] == "Query Id":
            header = [c if c is not None else "" for c in row]
            try:
                cols = {name: header.index(name) for name in REQUIRED_COLS}
            except ValueError as e:
                raise BadExport(f"Missing expected column: {e}")
            return buffer[:-1], header, cols
        if len(buffer) >= 30:
            break
    raise BadExport("Could not find the 'Query Id' header row — is this a Bulk Mentions export?")


# ---------------------------------------------------------------------------
# Writing — every output path (fill, enrichment) funnels through here so the
# xlsxwriter setup (constant memory, no auto-hyperlinks, date formatting)
# only needs to be right in one place.
# ---------------------------------------------------------------------------

def _write_cell(ws, r, c, val, date_fmt):
    if val is None or val == "":
        return
    if isinstance(val, bool):
        ws.write_boolean(r, c, val)
    elif isinstance(val, (int, float)):
        ws.write_number(r, c, val)
    elif isinstance(val, (datetime.datetime, datetime.date)):
        ws.write_datetime(r, c, val, date_fmt)
    else:
        ws.write_string(r, c, ILLEGAL.sub("", str(val)))


def _write_sheet(ws, rows, date_fmt):
    for r, row in enumerate(rows):
        for c, val in enumerate(row):
            _write_cell(ws, r, c, val, date_fmt)


def _safe_sheet_name(name):
    return (re.sub(r'[\[\]:*?/\\]', "-", name) or "Sheet")[:31]


def _write_workbook_bytes(main_rows, extra_sheets=None, extra_sheets_fn=None):
    """main_rows: any iterable of rows (list OR generator — constant_memory
    mode writes as it goes either way), row 0 is the header. extra_sheets:
    optional list of (sheet_name, rows) pairs, same shape. extra_sheets_fn:
    optional zero-arg callable returning more (sheet_name, rows) pairs,
    invoked AFTER main_rows is fully written but before the workbook closes
    — for sheets (e.g. Author Rollup) whose data is only known once a
    streamed main_rows generator has actually been drained."""
    buf = io.BytesIO()
    # NOT in_memory=True: xlsxwriter silently disables constant_memory when
    # in_memory is set ("We can't do 'constant_memory' mode while doing
    # 'in_memory' mode" — see its own workbook.py). Every write before this
    # fix was therefore running in full in-memory mode despite asking for
    # constant_memory — harmless for a 73k-row file (33MB output), but on a
    # 203k-row file it held the entire uncompressed sheet XML in RAM and
    # pushed peak RSS to ~2.9GB. Leaving in_memory unset routes constant_memory
    # mode's per-sheet buffering through real temp files instead, which is
    # exactly what keeps this flat — confirmed it still writes correctly to
    # a BytesIO target either way, since in_memory only controls where
    # constant_memory's intermediate sub-files live, not the final target.
    wb = xlsxwriter.Workbook(buf, {"constant_memory": True, "strings_to_urls": False})
    date_fmt = wb.add_format({"num_format": "yyyy-mm-dd hh:mm:ss"})
    _write_sheet(wb.add_worksheet("Sheet0"), main_rows, date_fmt)
    resolved_extras = list(extra_sheets or [])
    if extra_sheets_fn:
        resolved_extras += list(extra_sheets_fn())
    for name, rows in resolved_extras:
        _write_sheet(wb.add_worksheet(_safe_sheet_name(name)), rows, date_fmt)
    wb.close()
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Initial parse (upload -> URL list + full row data, ready for filling)
# ---------------------------------------------------------------------------

class ParsedExport:
    """Holds everything needed to both estimate (count URLs) and later fill/
    write. For a normal-size file, data_rows holds every row so fill_export
    never has to re-read the file. For a file over STREAMING_SIZE_THRESHOLD
    (is_streaming=True), data_rows is None — fill_export re-reads file_bytes
    itself, streaming, to keep peak memory flat regardless of file size."""

    def __init__(self, is_csv, urls, *, preamble, header, data_rows, cols, is_streaming=False):
        self.is_csv = is_csv
        self.urls = urls
        self.preamble = preamble
        self.header = header
        self.data_rows = data_rows
        self.cols = cols
        self.is_streaming = is_streaming


def parse_export(file_bytes, filename):
    """Parse an uploaded Bulk Mentions export: locate the header, extract
    every data row (not just URLs) so fill_export never has to re-read the
    file from scratch — unless the file is big enough that holding every
    row in memory risks OOMing the hosted app, in which case only the URLs
    are extracted here (streaming pass 1) and fill_export re-streams the
    file itself for pass 2."""
    is_csv = filename.lower().endswith(".csv")

    if _predict_needs_streaming(file_bytes, is_csv):
        row_iter = _stream_csv_rows(file_bytes) if is_csv else _stream_xlsx_rows(file_bytes)
        preamble, header, cols = _locate_header_streaming(row_iter)
        i_url = cols["Url"]
        urls = [str(row[i_url]).strip() for row in row_iter if len(row) > i_url and row[i_url]]
        if not urls:
            raise BadExport("No URLs found — is this a Bulk Mentions export?")
        return ParsedExport(is_csv, urls, preamble=preamble, header=header, data_rows=None, cols=cols,
                             is_streaming=True)

    rows = _read_csv_rows(file_bytes) if is_csv else _read_xlsx_rows(file_bytes)
    hdr_i, cols = locate_header(rows)
    preamble, header, data_rows = rows[:hdr_i], rows[hdr_i], rows[hdr_i + 1:]
    header = [c if c is not None else "" for c in header]
    data_rows = [list(row) for row in data_rows]

    urls = [str(r[cols["Url"]]).strip() for r in data_rows if len(r) > cols["Url"] and r[cols["Url"]]]
    if not urls:
        raise BadExport("No URLs found — is this a Bulk Mentions export?")
    return ParsedExport(is_csv, urls, preamble=preamble, header=header, data_rows=data_rows, cols=cols)


def read_header(file_bytes, filename):
    """Return the full header row (every column name, not just the required
    5) of a Bulk Mentions export — csv or xlsx, filled or not. Used to
    detect which columns an earlier module (in this pass or a prior one)
    already added, e.g. Sentiment Coding's "LLM Sentiment: <entity>" columns."""
    is_csv = filename.lower().endswith(".csv")
    if _predict_needs_streaming(file_bytes, is_csv):
        row_iter = _stream_csv_rows(file_bytes) if is_csv else _stream_xlsx_rows(file_bytes)
        _, header, _ = _locate_header_streaming(row_iter)
        return header
    rows = _read_csv_rows(file_bytes) if is_csv else _read_xlsx_rows(file_bytes)
    hdr_i, _ = locate_header(rows)
    return [c if c is not None else "" for c in rows[hdr_i]]


def iter_column_values(file_bytes, filename, column_name):
    """Read one named column's raw values across every data row of a Bulk
    Mentions export (csv or xlsx, filled or not) — for a lightweight preview/
    discovery step that doesn't need the full parse/fill machinery. Returns
    [] if the column isn't present in this export."""
    is_csv = filename.lower().endswith(".csv")
    if _predict_needs_streaming(file_bytes, is_csv):
        row_iter = _stream_csv_rows(file_bytes) if is_csv else _stream_xlsx_rows(file_bytes)
        _, header, _ = _locate_header_streaming(row_iter)
        if column_name not in header:
            return []
        idx = header.index(column_name)
        return [r[idx] if len(r) > idx else "" for r in row_iter]
    rows = _read_csv_rows(file_bytes) if is_csv else _read_xlsx_rows(file_bytes)
    hdr_i, _ = locate_header(rows)
    header = [c if c is not None else "" for c in rows[hdr_i]]
    if column_name not in header:
        return []
    idx = header.index(column_name)
    return [r[idx] if len(r) > idx else "" for r in rows[hdr_i + 1:]]


def _build_rollup_rows(author_stats):
    rows = [["Author", "Total Mentions", "Posts", "Comments", "Total Score", "Avg Score", "First Date", "Last Date"]]
    for author, a in sorted(author_stats.items(), key=lambda kv: kv[1]["count"], reverse=True):
        avg = round(a["score_sum"] / a["score_n"], 1) if a["score_n"] else ""
        rows.append([
            author, a["count"], a["posts"], a["comments"],
            a["score_sum"] if a["score_n"] else "",
            avg,
            a["first"].strftime("%Y-%m-%d") if a["first"] else "",
            a["last"].strftime("%Y-%m-%d") if a["last"] else "",
        ])
    return rows


class _FillState:
    """Mutable accumulator threaded through _make_row_transformer, shared by
    both the in-memory (list comprehension) and streaming (generator) fill
    paths so their row-transform behavior — and this bookkeeping — is
    provably identical."""

    def __init__(self):
        self.filled = 0
        self.truncated = 0
        self.unmatched = []  # (row_num, url, reason)
        self.author_stats = defaultdict(lambda: {"count": 0, "posts": 0, "comments": 0,
                                                   "score_sum": 0, "score_n": 0, "first": None, "last": None})


def _make_row_transformer(state, cols, preamble_len, by_url):
    """Returns transform(idx, row) -> output row, updating `state` as a side
    effect. Built once per fill_export call (not per row) since the closures
    only depend on cols/preamble_len/by_url, all fixed for the whole run."""
    i_date, i_url, i_ft = cols["Date"], cols["Url"], cols["Full Text"]
    i_author, i_title = cols.get("Author"), cols.get("Title")

    def lookup(url):
        return by_url.get(norm(url))

    def matched(r):
        return r if r and (r.get("status") or "").startswith("OK") else None

    def render_text(r):
        text = md_to_text(ILLEGAL.sub("", r.get("text") or ""))
        if len(text) > EXCEL_CELL_LIMIT:
            text = text[:EXCEL_CELL_LIMIT - 15] + " [...TRUNCATED]"
            state.truncated += 1
        return text

    def render_title(r):
        post_title = md_to_text(ILLEGAL.sub("", r.get("post_title") or ""))
        if len(post_title) > EXCEL_CELL_LIMIT:
            post_title = post_title[:EXCEL_CELL_LIMIT - 15] + " [...TRUNCATED]"
            state.truncated += 1
        return post_title

    def parse_author(r):
        return md_to_text(ILLEGAL.sub("", r.get("author") or ""))

    def parse_dt(r):
        if not r.get("created"):
            return None
        return datetime.datetime.fromisoformat(r["created"].replace("Z", "+00:00")).replace(tzinfo=None)

    def update_rollup(r, dt):
        author = r.get("author") or "[unknown/deleted]"
        a = state.author_stats[author]
        a["count"] += 1
        t = row_type(r)
        if t == "Post":
            a["posts"] += 1
        elif t == "Comment":
            a["comments"] += 1
        score = parse_score(r.get("score"))
        if score is not None:
            a["score_sum"] += score
            a["score_n"] += 1
        if dt:
            if a["first"] is None or dt < a["first"]:
                a["first"] = dt
            if a["last"] is None or dt > a["last"]:
                a["last"] = dt

    def transform(idx, row):
        if len(row) <= i_ft:
            return list(row) + [""] * len(EXTRA_COLS)
        url = str(row[i_url]).strip() if row[i_url] is not None else ""
        r = lookup(url) if url else None
        base = [ILLEGAL.sub("", c) if isinstance(c, str) else c for c in row]
        m = matched(r)
        if not m:
            if url:
                state.unmatched.append((preamble_len + 2 + idx, url, categorize_status(r)))
            return base + list(build_extra_cols(r).values())
        out = list(base)
        out[i_ft] = render_text(m)
        if i_author is not None and len(out) > i_author:
            out[i_author] = parse_author(m)
        if i_title is not None and len(out) > i_title:
            out[i_title] = render_title(m)
        dt = parse_dt(m)
        if dt:
            out[i_date] = dt
        result = out + list(build_extra_cols(m).values())
        update_rollup(m, dt)
        state.filled += 1
        return result

    return transform


def _stream_transform_rows(file_bytes, parsed, transform):
    """Pass 2 of the streaming fill: re-open file_bytes fresh, skip past the
    preamble + header rows already consumed in pass 1 (parse_export), then
    yield each transformed data row lazily."""
    row_source = _stream_csv_rows(file_bytes) if parsed.is_csv else _stream_xlsx_rows(file_bytes)
    it = iter(row_source)
    for _ in range(len(parsed.preamble) + 1):
        next(it, None)
    for idx, row in enumerate(it):
        yield transform(idx, row)


def fill_export(parsed: ParsedExport, fetch_results, file_bytes=None):
    """Fill Date/Full Text/metadata into the export and return (output_bytes, stats).

    fetch_results must be in the same order as parsed.urls (output of
    reddit_fetch.fetch_archive). file_bytes is required (and re-streamed)
    only when parsed.is_streaming — the normal path already has every row
    in parsed.data_rows and never touches it.
    """
    by_url = {norm(r["url"]): r for r in fetch_results}
    state = _FillState()
    transform = _make_row_transformer(state, parsed.cols, len(parsed.preamble), by_url)
    header_row = list(parsed.header) + EXTRA_COLS

    if parsed.is_streaming:
        if file_bytes is None:
            raise BadExport("Internal error: a streaming-mode parse requires file_bytes to fill.")
        main_rows = itertools.chain(parsed.preamble, [header_row],
                                     _stream_transform_rows(file_bytes, parsed, transform))
    else:
        out_rows = [transform(idx, row) for idx, row in enumerate(parsed.data_rows)]
        main_rows = list(parsed.preamble) + [header_row] + out_rows

    output_bytes = _write_workbook_bytes(
        main_rows,
        extra_sheets_fn=lambda: [("Author Rollup", _build_rollup_rows(state.author_stats))],
    )

    stats = {
        "filled": state.filled,
        "total": len(parsed.urls),
        "truncated": state.truncated,
        "unmatched": state.unmatched,
        "author_count": len(state.author_stats),
    }
    return io.BytesIO(output_bytes), stats


# ---------------------------------------------------------------------------
# Post-fill enrichment (Sentiment / Geolocation / Theme Summary / Driver
# Analysis) — read the already-filled sheet, append columns, write once.
# ---------------------------------------------------------------------------

class Sheet:
    """An in-memory, already-filled Bulk Mentions sheet that a module reads
    from and appends columns to. Plain Python lists throughout — see module
    docstring for why this replaced a live openpyxl worksheet. Unlike
    Mention Filler, this still needs the whole matrix resident (random-access
    column writes, and some modules need a full-corpus view) — see module
    docstring for why that's a real, separate size ceiling from Mention
    Filler's, not yet solved by streaming."""

    def __init__(self, preamble, header, rows):
        self.preamble = preamble
        self.header = list(header)
        self.rows = rows  # list of lists, mutable, parallel to header
        self._reindex()

    def _reindex(self):
        self.col_index = {name: i for i, name in enumerate(self.header) if name}

    def ensure_columns(self, names):
        added = [n for n in names if n not in self.col_index]
        for n in added:
            self.col_index[n] = len(self.header)
            self.header.append(n)
        if added:
            pad = [""] * len(added)
            for row in self.rows:
                row.extend(pad)

    def row_dict(self, row):
        return {name: (row[i] if i < len(row) else None) for name, i in self.col_index.items()}

    def set_by_index(self, row_idx, name, value):
        self.rows[row_idx][self.col_index[name]] = value

    def iter_rows(self):
        """Yield (row_index, {column_name: value}) for every data row."""
        for i, row in enumerate(self.rows):
            yield i, self.row_dict(row)

    def to_bytes(self, extra_sheets=None):
        """extra_sheets: optional list of (sheet_name, header_row, data_rows)."""
        main_rows = list(self.preamble) + [self.header] + self.rows
        extras = [(name, [header] + list(data)) for name, header, data in (extra_sheets or [])]
        return _write_workbook_bytes(main_rows, extra_sheets=extras)


def load_sheet_for_enrichment(file_bytes):
    """Load an already-filled xlsx (Mention Filler's output, or one the user
    uploaded pre-filled) into a Sheet for a module to read/append columns to.
    Uses the same size-triggered reader swap as parse_export for the read
    step (roughly halves peak memory on a huge file) but the Sheet itself
    still holds every row — see module docstring."""
    if _predict_needs_streaming(file_bytes, is_csv=False):
        rows = list(_stream_xlsx_rows(file_bytes))
    else:
        rows = _read_xlsx_rows(file_bytes)
    hdr_i = next((i for i, r in enumerate(rows[:30]) if r and r[0] == "Query Id"), None)
    if hdr_i is None:
        raise BadExport("Could not find the 'Query Id' header row — is this a filled Bulk Mentions export?")
    preamble, header, data_rows = rows[:hdr_i], rows[hdr_i], rows[hdr_i + 1:]
    header = [c if c is not None else "" for c in header]
    data_rows = [list(row) for row in data_rows]
    return Sheet(preamble, header, data_rows)


def unmatched_csv_bytes(unmatched):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Row", "Url", "Reason"])
    w.writerows(unmatched)
    return buf.getvalue().encode("utf-8")
